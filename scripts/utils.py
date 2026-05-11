# ============================================================================
# utils.py - Helper Functions for CAD VLM Extraction
# ============================================================================
"""
Utility functions for PDF processing, model management, and output generation.
"""

import os
import gc
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

import fitz  # PyMuPDF
from PIL import Image
import torch
import pandas as pd
from tqdm import tqdm


# ── File handling ──────────────────────────────────────────────────────────

def sanitize_filename(filename: str) -> str:
    """Remove invalid filename characters."""
    for char in '<>:"/\\|?*':
        filename = filename.replace(char, '_')
    return filename


def ensure_dir(path: Path) -> None:
    """Ensure directory exists."""
    path.mkdir(parents=True, exist_ok=True)


# ── PDF to Image conversion ────────────────────────────────────────────────

def pdf_to_image(
    pdf_path: str | Path,
    zoom: float = 2.0,
    max_size: Optional[int] = None
) -> Image.Image:
    """
    Convert first page of PDF to PIL Image.

    Args:
        pdf_path: Path to PDF file
        zoom: Zoom factor (1.0 = 72 DPI, 2.0 = 144 DPI, etc.)
        max_size: Max dimension in pixels (resizes if exceeded)

    Returns:
        PIL.Image object
    """
    doc = fitz.open(str(pdf_path))
    page = doc[0]

    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat)

    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()

    # Resize if too large
    if max_size and max(image.size) > max_size:
        ratio = max_size / max(image.size)
        new_size = (int(image.width * ratio), int(image.height * ratio))
        image = image.resize(new_size, Image.Resampling.LANCZOS)
        print(f"   Resized: {pix.width}x{pix.height} → {new_size[0]}x{new_size[1]}")
    else:
        print(f"   Image: {image.width}x{image.height} px ({zoom}x zoom)")

    return image


def pdf_to_image_path(
    pdf_path: str | Path,
    output_dir: str | Path,
    zoom: float = 2.0,
    max_size: Optional[int] = None
) -> str:
    """
    Convert PDF to image and save to disk.

    Args:
        pdf_path: Path to PDF file
        output_dir: Directory to save image
        zoom: Zoom factor
        max_size: Max dimension

    Returns:
        Path to saved image
    """
    image = pdf_to_image(pdf_path, zoom, max_size)

    output_dir = Path(output_dir)
    ensure_dir(output_dir)

    pdf_name = Path(pdf_path).stem
    image_path = output_dir / f"{pdf_name}.png"
    image.save(image_path, "PNG", optimize=True)

    return str(image_path)


# ── GPU memory management ───────────────────────────────────────────────────

def free_gpu_memory() -> None:
    """Free GPU memory."""
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()


def get_gpu_memory_info() -> Dict[str, Any]:
    """Get current GPU memory usage."""
    if not torch.cuda.is_available():
        return {"mode": "CPU", "allocated_gb": 0, "total_gb": 0}

    allocated = torch.cuda.memory_allocated(0) / 1e9
    reserved = torch.cuda.memory_reserved(0) / 1e9
    total = torch.cuda.get_device_properties(0).total_memory / 1e9

    return {
        "mode": "GPU",
        "device": torch.cuda.get_device_name(0),
        "allocated_gb": round(allocated, 2),
        "reserved_gb": round(reserved, 2),
        "total_gb": round(total, 2),
        "free_gb": round(total - allocated, 2)
    }


def can_load_model(vram_needed_gb: float, safety_margin_gb: float = 1.0) -> bool:
    """Check if model can fit in GPU memory."""
    info = get_gpu_memory_info()
    if info["mode"] == "CPU":
        return True  # Will run on CPU
    return info["free_gb"] >= (vram_needed_gb + safety_margin_gb)


# ── Save functions ──────────────────────────────────────────────────────────

def save_json(data: Dict, filepath: str | Path) -> None:
    """Save data to JSON file."""
    filepath = Path(filepath)
    ensure_dir(filepath.parent)

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"   Saved: {filepath.name}")


def save_excel(
    data: List[Dict],
    filepath: str | Path,
    sheet_name: str = "Results"
) -> None:
    """Save data to Excel with formatting."""
    filepath = Path(filepath)
    ensure_dir(filepath.parent)

    df = pd.DataFrame(data)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format header row
        ws = writer.sheets[sheet_name]

        from openpyxl.styles import Font, Alignment, PatternFill

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-adjust columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 80)

    print(f"   Saved: {filepath.name}")


def save_comparison_excel(
    results: Dict[str, Dict],
    filepath: str | Path
) -> None:
    """
    Save comparison results to multi-sheet Excel.

    Args:
        results: Dict[drawing_name, Dict[model_name, extraction_data]]
        filepath: Output Excel path
    """
    filepath = Path(filepath)
    ensure_dir(filepath.parent)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = []
        for drawing, models_data in results.items():
            row = {"Drawing": drawing}
            for model_name, data in models_data.items():
                row[f"{model_name}_Total"] = data.get("total_extracted", 0)
                row[f"{model_name}_Categories"] = data.get("categories_count", 0)
            summary_data.append(row)

        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # Per-drawing comparison sheets
        for drawing, models_data in results.items():
            sheet_name = sanitize_filename(drawing)[:30]  # Excel limit
            rows = []

            for category, params in results[drawing].get("categories", {}).items():
                for param, value in params.items():
                    rows.append({
                        "Category": category,
                        "Parameter": param,
                        "Value": value
                    })

            if rows:
                pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"   Saved comparison: {filepath.name}")


# ── OCR text categorization ──────────────────────────────────────────────────

def categorize_text_elements(elements: List[Dict]) -> Dict[str, List[Dict]]:
    """
    Categorize OCR text elements based on patterns.

    Args:
        elements: List of {text, confidence, x, y} from OCR

    Returns:
        Dict with categories and matched elements
    """
    categories = {
        "Dimensions_Tolerances": [],
        "GD_T": [],
        "Views_Drafting": [],
        "Notes_Annotations": [],
        "TitleBlock_Metadata": [],
        "Uncategorized": []
    }

    # Regex patterns
    patterns = {
        "diameter": [r'^[F\[\(]?\d+', r'^Ø?\s*\d+'],  # Diameter values
        "tolerance": [r'[+-]\d+\.\d+', r'[±+\-]\d+'],  # Tolerance values
        "radius": [r'^R\s*\d+', r'^R\d+'],  # Radius values
        "scale": [r'SCALE\s+\d+:\d+', r'\d+:\d+'],  # Scale notation
        "view": [r'(FRONT|TOP|SIDE|SECTION|DETAIL|A-A|B-B)', r'SCALE\s+\d+:\d+'],
        "gdt": [r'⊕', r'[A-Z]\s*[0-9]+\.[0-9]+'],  # GD&T symbols
        "title_block": [r'(REV|DATE|DRAWN|DESIGNED|CHECKED|APPROVED)', r'(PART|DWG|Drawing)'],
    }

    for elem in elements:
        text = elem.get("text", "")
        matched = False

        # Check diameter
        if re.search(r'[F\[\(]?\d+', text) or text.startswith('Ø'):
            categories["Dimensions_Tolerances"].append(elem)
            matched = True

        # Check tolerance
        elif re.search(r'[+-]\d+\.\d+', text):
            categories["Dimensions_Tolerances"].append(elem)
            matched = True

        # Check radius
        elif text.startswith('R') and re.search(r'\d', text):
            categories["Dimensions_Tolerances"].append(elem)
            matched = True

        # Check views
        elif re.search(r'(FRONT|TOP|SIDE|SECTION|DETAIL)', text.upper()):
            categories["Views_Drafting"].append(elem)
            matched = True

        # Check GD&T
        elif '⊕' in text or re.search(r'[A-Z]\s*[0-9]+\.[0-9]+', text):
            categories["GD_T"].append(elem)
            matched = True

        # Check title block keywords
        elif re.search(r'(REV|DATE|DRAWN|DESIGNED|CHECKED|APPROVED|PART)', text.upper()):
            categories["TitleBlock_Metadata"].append(elem)
            matched = True

        if not matched:
            # Check position - title block usually at bottom
            y = elem.get("y", 0)
            if y > 800:  # Lower part of drawing
                categories["TitleBlock_Metadata"].append(elem)
            else:
                categories["Uncategorized"].append(elem)

    return categories


# ── Post-processing for OCR results ────────────────────────────────────────

def fix_diameter_symbol(ocr_text: str) -> str:
    """
    Fix common misreadings of Ø symbol by EasyOCR.

    EasyOCR often misreads Ø as F, [, p, 0, etc.
    This function attempts to reconstruct diameter values.
    """
    # Common misreadings
    misreadings = [
        (r'\bF(\d+\.?\d*)', r'Ø\1'),  # F20 -> Ø20
        (r'\b\[(\d+\.?\d*)', r'Ø\1'),  # [20 -> Ø20
        (r'\b0(\d+\.?\d*)', r'Ø\1'),  # 020 -> Ø20
        (r'\bp(\d+\.?\d*)', r'Ø\1'),  # p20 -> Ø20
    ]

    result = ocr_text
    for pattern, replacement in misreadings:
        result = re.sub(pattern, replacement, result)

    return result


def merge_split_values(elements: List[Dict], threshold: float = 30) -> List[Dict]:
    """
    Merge split values that should be together.

    E.g., "609." and ".6" should merge to "609.6"
    """
    if not elements:
        return elements

    merged = []
    i = 0

    while i < len(elements):
        current = elements[i]
        current_text = str(current.get("text", ""))

        # Check if this could be start of split value
        if re.match(r'^\d+\.?$', current_text):  # Ends with number or decimal
            j = i + 1
            while j < len(elements):
                next_elem = elements[j]
                next_text = str(next_elem.get("text", ""))

                # Check if next starts with . or digit and close enough
                if re.match(r'^\.\d+', next_text):
                    y_diff = abs(current.get("y", 0) - next_elem.get("y", 0))
                    x_diff = abs(current.get("x", 0) - next_elem.get("x", 0))

                    if y_diff < threshold and x_diff < threshold * 2:
                        # Merge
                        current["text"] = current_text + next_text
                        current_text = current["text"]
                        j += 1
                    else:
                        break
                else:
                    break

        merged.append(current)
        i = j if j > i else i + 1

    return merged


# ── Progress display ────────────────────────────────────────────────────────

def print_separator(char: str = "=", length: int = 60) -> None:
    """Print a separator line."""
    print(char * length)


def print_header(title: str) -> None:
    """Print a formatted header."""
    print_separator()
    print(f"  {title}")
    print_separator()


def print_model_status(model_name: str, status: str, details: str = "") -> None:
    """Print model loading/testing status."""
    status_symbols = {
        "loading": "📦",
        "success": "✅",
        "error": "❌",
        "skipping": "⏭️",
    }
    symbol = status_symbols.get(status, "•")
    print(f"   {symbol} {model_name}: {details}")


# ── Main entry point check ──────────────────────────────────────────────────

if __name__ == "__main__":
    print("CAD VLM Extraction - Utilities Module")
    print(f"GPU Memory: {get_gpu_memory_info()}")