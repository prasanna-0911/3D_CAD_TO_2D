"""
utils.py - Helper Functions for 71 Parameters Extraction
============================================================
"""

import os
import gc
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

import fitz
from PIL import Image
import pandas as pd

# ============================================================================
# FILE HANDLING
# ============================================================================

def ensure_dir(path: Path) -> None:
    """Ensure directory exists."""
    path.mkdir(parents=True, exist_ok=True)

def sanitize_filename(filename: str) -> str:
    """Remove invalid filename characters."""
    for char in '<>:"/\\|?*':
        filename = filename.replace(char, '_')
    return filename

# ============================================================================
# PDF TO IMAGE CONVERSION
# ============================================================================

def pdf_to_image(pdf_path: str | Path, zoom: float = 2.0, max_size: Optional[int] = None) -> Image.Image:
    """Convert first page of PDF to PIL Image."""
    doc = fitz.open(str(pdf_path))
    page = doc[0]
    
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat)
    
    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()
    
    if max_size and max(image.size) > max_size:
        ratio = max_size / max(image.size)
        new_size = (int(image.width * ratio), int(image.height * ratio))
        image = image.resize(new_size, Image.Resampling.LANCZOS)
        print(f"   Resized: {pix.width}x{pix.height} -> {new_size[0]}x{new_size[1]}")
    else:
        print(f"   Image: {image.width}x{image.height} px ({zoom}x zoom)")
    
    return image

# ============================================================================
# SAVE FUNCTIONS
# ============================================================================

def save_json(data: Dict, filepath: str | Path) -> None:
    """Save data to JSON file."""
    filepath = Path(filepath)
    ensure_dir(filepath.parent)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    print(f"   Saved: {filepath.name}")

def save_excel(data: List[Dict], filepath: str | Path, sheet_name: str = "Results") -> None:
    """Save data to Excel with formatting."""
    filepath = Path(filepath)
    ensure_dir(filepath.parent)
    
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        ws = writer.sheets[sheet_name]
        
        from openpyxl.styles import Font, Alignment, PatternFill
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 80)
    
    print(f"   Saved: {filepath.name}")

# ============================================================================
# PARAMETER DETECTION FUNCTIONS
# ============================================================================

def check_71_param_detection(text: str, param_id: str, param_name: str) -> bool:
    """
    Check if a 71 parameter is detected in extracted text.
    
    Uses keyword matching for each parameter category.
    """
    text_lower = text.lower()
    param_lower = param_name.lower()
    
    # Category-specific detection logic
    if param_id in ["5", "6", "7", "8"]:  # Holes, Fillet, Chamfers, Ribs
        keywords = param_name.replace(" Feature", "").lower().split()
        return any(kw in text_lower for kw in keywords if len(kw) > 3)
    
    elif param_id in ["9", "10", "11", "12"]:  # Size, Length, Diameter, Thickness
        return any(kw in param_lower for kw in text_lower if len(kw) > 2)
    
    elif param_id in ["25", "26", "27", "28", "29", "30", "31", "32"]:  # GD&T
        gdt_symbols = {
            "25": ["position", "pos", "⊕"],
            "26": ["straightness", "straight", "—"],
            "27": ["flatness", "flat", "□"],
            "28": ["circularity", "circular", "○"],
            "29": ["parallelism", "parallel", "//"],
            "30": ["perpendicularity", "perpendicular", "⊥"],
            "31": ["angularity", "angular", "∠"],
            "32": ["runout", "run-out", "↗"]
        }
        keywords = gdt_symbols.get(param_id, [])
        return any(kw in text_lower for kw in keywords)
    
    elif param_id in ["33", "34", "35", "36", "45"]:  # Views
        view_keywords = ["front", "top", "side", "section", "detail", "view", "scale"]
        return any(kw in text_lower for kw in view_keywords)
    
    elif param_id in ["51", "52", "53", "54"]:  # Notes
        note_keywords = ["note", "warning", "caution", "note:", "1)", "2)"]
        return any(kw in text_lower for kw in note_keywords)
    
    elif param_id in ["59", "60", "61", "62", "63", "64", "65"]:  # TitleBlock
        title_keywords = ["rev", "revision", "dwg", "drawing", "part", "name", "date", "drawn", "checked"]
        return any(kw in text_lower for kw in title_keywords)
    
    elif param_id in ["66", "67", "68", "69"]:  # BOM
        bom_keywords = ["bom", "bill", "material", "parts", "item", "qty", "quantity"]
        return any(kw in text_lower for kw in bom_keywords)
    
    elif param_id in ["70", "71"]:  # Scale
        scale_keywords = ["scale", "1:", "1/"]
        return any(kw in text_lower for kw in scale_keywords)
    
    # General keyword matching
    keywords = param_lower.split()
    return any(kw in text_lower for kw in keywords if len(kw) > 3)

def detect_params_71(extracted_text: str) -> Dict[str, Any]:
    """
    Detect all 71 parameters in extracted text.
    
    Returns: Dict with param_id -> {name, detected, value}
    """
    from config import ALL_PARAMS_71
    
    detected_params = {}
    
    for param_id, param_info in ALL_PARAMS_71.items():
        is_detected = check_71_param_detection(
            extracted_text, 
            param_id, 
            param_info["name"]
        )
        
        detected_params[param_id] = {
            "id": param_id,
            "name": param_info["name"],
            "detected": is_detected,
            "category": next(cat for cat, params in __import__('config', fromlist=['PARAMETERS_71']).PARAMETERS_71.items() 
                        for p in params if p["id"] == param_id)
        }
    
    return detected_params

def detect_ground_truth_params(extracted_text: str, param_id: str) -> Optional[float]:
    """
    Try to detect ground truth parameter value from extracted text.
    
    Returns: Numeric value if found, None otherwise
    """
    import re
    
    # Mapping of GT params to regex patterns
    patterns = {
        "GT_10": r"(\d+)\s*views?",  # No_of_Views
        "GT_12": r"top.*?x[\s:=]*(\d+\.?\d*)",  # View_Top_X
        "GT_13": r"top.*?y[\s:=]*(\d+\.?\d*)",  # View_Top_Y
        "GT_17": r"front.*?x[\s:=]*(\d+\.?\d*)",  # View_Front_X
        "GT_18": r"front.*?y[\s:=]*(\d+\.?\d*)",  # View_Front_Y
    }
    
    pattern = patterns.get(param_id)
    if not pattern:
        return None
    
    match = re.search(pattern, extracted_text.lower())
    if match:
        try:
            return float(match.group(1))
        except:
            pass
    
    return None

# ============================================================================
# DISPLAY FUNCTIONS
# ============================================================================

def print_separator(char: str = "=", length: int = 60) -> None:
    """Print a separator line."""
    print(char * length)

def print_header(title: str) -> None:
    """Print a formatted header."""
    print_separator()
    print(f"  {title}")
    print_separator()

def print_model_status(model_name: str, status: str, details: str = "") -> None:
    """Print model loading/testing status."""
    status_symbols = {
        "loading": "[LOADING]",
        "success": "[OK]",
        "error": "[ERROR]",
        "skipping": "[SKIP]",
    }
    symbol = status_symbols.get(status, "•")
    print(f"   {symbol} {model_name}: {details}")

# ============================================================================
# GPU MEMORY MANAGEMENT
# ============================================================================

def free_gpu_memory() -> None:
    """Free GPU memory."""
    gc.collect()
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            torch.cuda.synchronize()
    except:
        pass

def get_gpu_memory_info() -> Dict[str, Any]:
    """Get current GPU memory usage."""
    try:
        import torch
        if not torch.cuda.is_available():
            return {"mode": "CPU", "available": False, "allocated_gb": 0, "total_gb": 0, "free_gb": 0}
        
        allocated = torch.cuda.memory_allocated(0) / 1e9
        total = torch.cuda.get_device_properties(0).total_memory / 1e9
        
        return {
            "mode": "GPU",
            "available": True,
            "device": torch.cuda.get_device_name(0),
            "allocated_gb": round(allocated, 2),
            "total_gb": round(total, 2),
            "free_gb": round(total - allocated, 2)
        }
    except:
        return {"mode": "CPU", "available": False, "allocated_gb": 0, "total_gb": 0, "free_gb": 0}

def can_load_model(vram_needed_gb: float, safety_margin_gb: float = 1.0) -> bool:
    """Check if model can fit in GPU memory."""
    info = get_gpu_memory_info()
    if info["mode"] == "CPU":
        return True
    return info["free_gb"] >= (vram_needed_gb + safety_margin_gb)

# ============================================================================
# COMPARISON FUNCTIONS
# ============================================================================

def compare_with_ground_truth(model_value: Any, ground_truth_value: Any) -> Dict[str, Any]:
    """
    Compare model extracted value with ground truth.
    
    Returns: {match: bool, difference: float, percentage_diff: float}
    """
    try:
        model_val = float(model_value)
        gt_val = float(ground_truth_value)
        
        if gt_val == 0:
            is_match = model_val == 0
            diff = abs(model_val - gt_val)
        else:
            diff = abs(model_val - gt_val)
            pct_diff = (diff / abs(gt_val)) * 100
            is_match = pct_diff < 10  # Within 10% considered match
        
        return {
            "match": is_match,
            "difference": round(diff, 4),
            "percentage_diff": round(pct_diff, 2) if gt_val != 0 else None,
            "model_value": model_val,
            "ground_truth": gt_val
        }
    except:
        return {
            "match": False,
            "difference": None,
            "percentage_diff": None,
            "model_value": str(model_value),
            "ground_truth": str(ground_truth_value)
        }

if __name__ == "__main__":
    print("71 Parameters Extraction - Utilities Module")