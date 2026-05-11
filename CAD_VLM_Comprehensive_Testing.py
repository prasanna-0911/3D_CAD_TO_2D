# ============================================================================
# CAD_VLM_Comprehensive_Testing.py
# ============================================================================
"""
Complete CAD Drawing VLM Testing Framework for Google Colab
============================================================
Tests all 71 parameters across 5 models with JSON + Excel output.

Models: EasyOCR, LLaVA 1.5, BLIP2, Moondream2, Qwen2-VL
Output: Google Drive (JSON + Excel for each drawing)
"""

# ============================================================================
# CELL 0 - MOUNT GOOGLE DRIVE
# ============================================================================
from google.colab import drive
drive.mount('/content/drive')

import os
os.chdir("/content/drive/MyDrive/3D_CAD_TO_2D")

print(f"Working directory: {os.getcwd()}")
print("Google Drive mounted successfully!")

# ============================================================================
# CELL 1 - INSTALL DEPENDENCIES
# ============================================================================
!pip install -q torch torchvision torchaudio
!pip install -q transformers accelerate bitsandbytes
!pip install -q easyocr pymupdf pandas openpyxl tqdm
!pip install -q qwen-vl-utils

print("✅ All dependencies installed!")

# ============================================================================
# CELL 2 - IMPORT LIBRARIES
# ============================================================================
import os
import gc
import json
import torch
from pathlib import Path
from datetime import datetime
from tqdm import tqdm
import pandas as pd

print("✅ Libraries imported!")

# ============================================================================
# CELL 3 - DEFINE 71 PARAMETERS STRUCTURE
# ============================================================================
PARAMETERS = {
    "Geometry_Design": [
        {"id": 1, "name": "Geometry & Design Changes", "description": "Overall geometry modifications"},
        {"id": 2, "name": "Shape modifications", "description": "Profiles, contours changes"},
        {"id": 3, "name": "Suppressed/unsuppressed features", "description": "Feature visibility changes"},
        {"id": 4, "name": "Coordinate system shifts", "description": "CSYS location changes"},
        {"id": 5, "name": "Holes Feature", "description": "Additions/removals of holes"},
        {"id": 6, "name": "Fillet Feature", "description": "Additions/removals of fillets"},
        {"id": 7, "name": "Chamfers Feature", "description": "Additions/removals of chamfers"},
        {"id": 8, "name": "Ribs Feature", "description": "Additions/removals of ribs"},
    ],

    "Dimensions": [
        {"id": 9, "name": "Size changes", "description": "Overall size modifications"},
        {"id": 10, "name": "Length", "description": "Linear dimension values"},
        {"id": 11, "name": "Diameter", "description": "Diameter dimensions (Ø)"},
        {"id": 12, "name": "Thickness", "description": "Wall thickness dimensions"},
        {"id": 13, "name": "Position changes", "description": "Feature position modifications"},
        {"id": 14, "name": "Location changes of features", "description": "Feature spacing changes"},
        {"id": 19, "name": "Dimension position changes", "description": "Dimension placement changes"},
        {"id": 22, "name": "Dimension value changes", "description": "Numerically different dimensions"},
    ],

    "Assembly": [
        {"id": 15, "name": "Assembly fit/interface changes", "description": "Fit tolerance modifications"},
        {"id": 16, "name": "Exploded view differences", "description": "Explosion view changes"},
        {"id": 17, "name": "Sub-assembly changes", "description": "Sub-assembly modifications"},
        {"id": 18, "name": "Fastener type/size updates", "description": "Bolt, screw changes"},
    ],

    "Tolerances": [
        {"id": 20, "name": "Tolerances change", "description": "General tolerance modifications"},
        {"id": 21, "name": "Tolerances Location change", "description": "Tolerance placement changes"},
        {"id": 23, "name": "Tolerance updates", "description": "Bilateral/unilateral changes"},
        {"id": 24, "name": "Fits", "description": "H7/g6 type fit specifications"},
    ],

    "GD_T": [
        {"id": 25, "name": "Position tolerance", "description": "Positional GD&T symbol"},
        {"id": 26, "name": "Straightness", "description": "Straightness GD&T"},
        {"id": 27, "name": "Flatness", "description": "Flatness GD&T"},
        {"id": 28, "name": "Circularity", "description": "Circularity GD&T"},
        {"id": 29, "name": "Parallelism", "description": "Parallelism GD&T"},
        {"id": 30, "name": "Perpendicularity", "description": "Perpendicularity GD&T"},
        {"id": 31, "name": "Angularity", "description": "Angularity GD&T"},
        {"id": 32, "name": "Runout", "description": "Runout GD&T"},
    ],

    "Views_Drafting": [
        {"id": 33, "name": "View changes", "description": "View type modifications"},
        {"id": 34, "name": "View add/delete", "description": "New or removed views"},
        {"id": 35, "name": "View scale change", "description": "Individual view scale changes"},
        {"id": 36, "name": "View representation change", "description": "View style modifications"},
        {"id": 37, "name": "Drafting view", "description": "Drafting view specifications"},
        {"id": 38, "name": "Section line", "description": "Section line specifications"},
        {"id": 39, "name": "Hatching change", "description": "Section hatching modifications"},
        {"id": 40, "name": "Line type change", "description": "Line style modifications"},
        {"id": 41, "name": "Fake dimension", "description": "Reference dimension changes"},
        {"id": 42, "name": "Dimension arrow change", "description": "Arrow style modifications"},
        {"id": 43, "name": "Dimension line thickness change", "description": "Line weight changes"},
        {"id": 44, "name": "Dimension position change", "description": "Dimension placement shifts"},
        {"id": 45, "name": "View position changes", "description": "View layout changes"},
        {"id": 46, "name": "Leader line change", "description": "Leader line modifications"},
        {"id": 47, "name": "Additional curve creation", "description": "Curve additions for dimensioning"},
        {"id": 48, "name": "Section view shift", "description": "Section view position changes"},
        {"id": 49, "name": "Broken view shift", "description": "Broken view modifications"},
        {"id": 50, "name": "Additional import photo/graphics", "description": "Image imports"},
    ],

    "Notes_Annotations": [
        {"id": 51, "name": "General notes updates", "description": "General note modifications"},
        {"id": 52, "name": "Notes Location", "description": "Note placement changes"},
        {"id": 53, "name": "Special instructions", "description": "Special instruction additions/removals"},
        {"id": 54, "name": "Flag notes", "description": "Key characteristic flags"},
        {"id": 55, "name": "Special symbols", "description": "Special symbol additions/deletions"},
        {"id": 56, "name": "Welding symbols", "description": "Welding symbol specifications"},
        {"id": 57, "name": "Datum changes", "description": "Datum reference modifications"},
        {"id": 58, "name": "Datum Location changes", "description": "Datum label position changes"},
    ],

    "TitleBlock_Metadata": [
        {"id": 59, "name": "Revision number", "description": "REV field value"},
        {"id": 60, "name": "Drawing number", "description": "Drawing ID number"},
        {"id": 61, "name": "Part name changes", "description": "Part title modifications"},
        {"id": 62, "name": "Author", "description": "Drawn by field"},
        {"id": 63, "name": "Checker/approver updates", "description": "Checked/Approved fields"},
        {"id": 64, "name": "Dates", "description": "Date field values"},
        {"id": 65, "name": "Company or project info", "description": "Company/project name"},
    ],

    "BOM": [
        {"id": 66, "name": "BOM", "description": "Bill of Materials presence"},
        {"id": 67, "name": "Component changes", "description": "Added/removed components"},
        {"id": 68, "name": "Part number revisions", "description": "Part number modifications"},
        {"id": 69, "name": "Part list", "description": "Part list specifications"},
    ],

    "Scale": [
        {"id": 70, "name": "Sheet Scale changes", "description": "Overall drawing scale"},
        {"id": 71, "name": "View Scale changes", "description": "Individual view scales"},
    ]
}

# Count total parameters
TOTAL_PARAMS = sum(len(params) for params in PARAMETERS.values())
print(f"Total parameters defined: {TOTAL_PARAMS}")

# ============================================================================
# CELL 4 - HELPER FUNCTIONS
# ============================================================================
import fitz  # PyMuPDF
from PIL import Image

def pdf_to_image(pdf_path, zoom=2.0, max_size=None):
    """Convert PDF to PIL Image."""
    doc = fitz.open(pdf_path)
    page = doc[0]
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat)
    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    doc.close()

    if max_size and max(image.size) > max_size:
        ratio = max_size / max(image.size)
        new_size = (int(image.width * ratio), int(image.height * ratio))
        image = image.resize(new_size, Image.Resampling.LANCZOS)

    return image


def save_json(data, filepath):
    """Save data to JSON file."""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def save_excel(results_df, filepath, sheet_name="Results"):
    """Save DataFrame to Excel with formatting."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 4, 60)


def create_results_dataframe(extraction_results, model_name):
    """Create a structured DataFrame from extraction results."""
    rows = []

    for category, params in PARAMETERS.items():
        for param in params:
            param_id = param["id"]
            param_name = param["name"]

            # Get extraction result for this parameter
            result = extraction_results.get(category, {}).get(str(param_id), {})

            rows.append({
                "Category": category,
                "Parameter_ID": param_id,
                "Parameter_Name": param_name,
                "Parameter_Description": param["description"],
                "Detected": result.get("detected", False),
                "Value": result.get("value", "Not found"),
                "Confidence": result.get("confidence", "N/A"),
                "Model": model_name
            })

    return pd.DataFrame(rows)


# ============================================================================
# CELL 5 - PDF LIST
# ============================================================================
# Find all PDFs
pdf_files = [
    "/content/drive/MyDrive/3D_CAD_TO_2D/input_pdfs/HP_V1__Final Shaft_2d.pdf",
    "/content/drive/MyDrive/3D_CAD_TO_2D/input_pdfs/HP_V2__Final Shaft_2d.pdf",
    "/content/drive/MyDrive/3D_CAD_TO_2D/input_pdfs/HP_V1__bolt.pdf",
    "/content/drive/MyDrive/3D_CAD_TO_2D/input_pdfs/HP_V2__bolt_2d.pdf",
    "/content/drive/MyDrive/3D_CAD_TO_2D/input_pdfs/HP_Before__1st__NOVEX-GS 414 UFS DA-1__2d.pdf",
    "/content/drive/MyDrive/3D_CAD_TO_2D/input_pdfs/HP_After__1st__NOVEX-GS 414 UFS DA-1__2d.pdf",
]

# Verify files exist
existing_pdfs = [f for f in pdf_files if os.path.exists(f)]
print(f"Found {len(existing_pdfs)} PDFs:")
for pdf in existing_pdfs:
    print(f"  - {os.path.basename(pdf)}")